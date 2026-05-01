#!/usr/bin/env python3
"""Converte uma planilha SurveyXLSX em SurveyMD.

Uso:
    python xlsx2md.py questionario_modelo.xlsx questionario.md

Dependência:
    pip install openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Instale a dependência com: pip install openpyxl") from exc

TRUE_VALUES = {"y", "yes", "true", "1", "sim", "s", "x"}
FALSE_VALUES = {"n", "no", "false", "0", "não", "nao", ""}

SURVEY_KEYS_ORDER = [
    "title",
    "language",
    "sid",
    "admin",
    "adminemail",
    "format",
    "template",
    "expires",
]

QUESTION_OPTION_KEYS = [
    "mandatory",
    "scale",
    "visible_if",
    "help",
    "evidence",
    "evidence_mandatory",
    "evidence_allowed_filetypes",
    "evidence_min_files",
    "evidence_max_files",
    "detail",
    "detail_mandatory",
    "min_answers",
    "max_answers",
    "allowed_filetypes",
    "min_files",
    "max_files",
    "default",
    "hide_tip",
]

SUBQUESTION_EVIDENCE_KEYS = [
    "evidence_mandatory",
    "evidence_allowed_filetypes",
    "evidence_min_files",
    "evidence_max_files",
    "evidence_text",
]

ORG_SPLIT_RE = re.compile(r"[,;|]")
SID_BY_ORG_RE = re.compile(r"^sid_(.+)$", re.IGNORECASE)
VISIBLE_IF_REF_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9_]*)\b(?=\s*(?:[.=]|==|!=|>=|<=|>|<))")


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def slug_bool(value: Any) -> bool:
    text = clean(value).lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return bool(text)


def sheet_rows(wb, sheet_name: str) -> List[Dict[str, str]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    expected_first_header = {
        "survey": "key",
        "scales": "scale_code",
        "groups": "group_code",
        "subgroups": "group_code",
        "questions": "group_code",
        "subquestions": "question_code",
        "options": "question_code",
    }.get(sheet_name)
    header_index = 0
    if expected_first_header:
        for idx, row in enumerate(rows):
            normalized = [clean(cell) for cell in row]
            if expected_first_header in normalized:
                header_index = idx
                break
        else:
            return []
    headers = [clean(h) for h in rows[header_index]]
    data: List[Dict[str, str]] = []
    for row in rows[header_index + 1:]:
        item = {headers[i]: clean(row[i]) if i < len(row) else "" for i in range(len(headers)) if headers[i]}
        if any(v for v in item.values()):
            data.append(item)
    return data


def survey_meta(wb) -> Dict[str, str]:
    meta: Dict[str, str] = {}
    for row in sheet_rows(wb, "survey"):
        key = row.get("key", "")
        if key:
            meta[key] = row.get("value", "")
    return meta


def parse_orgs(value: Any) -> List[str]:
    text = clean(value)
    if not text:
        return []
    return [part.strip() for part in ORG_SPLIT_RE.split(text) if part.strip()]


def row_applies_to_org(row: Dict[str, str], org: Optional[str]) -> bool:
    if not org:
        return True
    orgs = parse_orgs(row.get("orgs", ""))
    if not orgs:
        return True
    wanted = org.casefold()
    return any(item.casefold() == wanted for item in orgs)


def discover_organizations_from_rows(*row_sets: Iterable[Dict[str, str]]) -> List[str]:
    seen: Set[str] = set()
    orgs: List[str] = []
    for rows in row_sets:
        for row in rows:
            for org in parse_orgs(row.get("orgs", "")):
                key = org.casefold()
                if key not in seen:
                    seen.add(key)
                    orgs.append(org)
    return sorted(orgs, key=str.casefold)


def discover_organizations(input_path: Path) -> List[str]:
    wb = load_workbook(input_path, data_only=False)
    return discover_organizations_from_rows(
        sheet_rows(wb, "questions"),
        sheet_rows(wb, "subquestions"),
        sheet_rows(wb, "options"),
    )


def meta_for_org(meta: Dict[str, str], org: Optional[str]) -> Dict[str, str]:
    clean_meta = {key: value for key, value in meta.items() if not SID_BY_ORG_RE.match(key)}
    if org:
        for key, value in meta.items():
            match = SID_BY_ORG_RE.match(key)
            if match and match.group(1).casefold() == org.casefold():
                clean_meta["sid"] = value
                break
        else:
            clean_meta.pop("sid", None)
    return clean_meta


def filter_rows_for_org(
    questions: List[Dict[str, str]],
    subquestions: List[Dict[str, str]],
    options: List[Dict[str, str]],
    org: Optional[str],
) -> tuple[List[Dict[str, str]], List[Dict[str, str]], List[Dict[str, str]]]:
    if not org:
        return questions, subquestions, options

    filtered_questions = [row for row in questions if row_applies_to_org(row, org)]
    question_codes = {row.get("code", "") for row in filtered_questions if row.get("code", "")}
    filtered_subquestions = [
        row
        for row in subquestions
        if row.get("question_code", "") in question_codes and row_applies_to_org(row, org)
    ]
    filtered_options = [
        row
        for row in options
        if row.get("question_code", "") in question_codes and row_applies_to_org(row, org)
    ]
    return filtered_questions, filtered_subquestions, filtered_options


def validate_visible_if_references(
    questions: List[Dict[str, str]],
    org: Optional[str],
    original_question_codes: Set[str],
) -> None:
    codes = {row.get("code", "") for row in questions if row.get("code", "")}
    ignored_tokens = {"and", "or", "not"}
    for row in questions:
        code = row.get("code", "")
        visible_if = row.get("visible_if", "")
        if not visible_if:
            continue
        refs = [ref for ref in VISIBLE_IF_REF_RE.findall(visible_if) if ref.lower() not in ignored_tokens]
        removed = [ref for ref in refs if ref in original_question_codes and ref not in codes]
        if removed:
            suffix = f" para org '{org}'" if org else ""
            raise ValueError(
                f"Questao {code} tem visible_if que referencia questao removida ou inexistente{suffix}: "
                f"{', '.join(sorted(set(removed)))}"
            )


def sort_key(row: Dict[str, str]) -> tuple:
    raw = row.get("order", "") or row.get("parent_order", "")
    try:
        return (0, float(raw))
    except ValueError:
        return (1, raw)


def emit_frontmatter(meta: Dict[str, str]) -> List[str]:
    lines = ["---"]
    keys = [k for k in SURVEY_KEYS_ORDER if meta.get(k)]
    keys += sorted(k for k in meta.keys() if k not in keys and meta.get(k))
    for key in keys:
        value = meta[key]
        if "\n" in value:
            lines.append(f"{key}: |")
            for part in value.splitlines():
                lines.append(f"  {part}")
        else:
            escaped = value.replace('"', '\\"')
            if re.search(r"[:#{}\[\],&*?\-|<>=!%@`]", value):
                lines.append(f'{key}: "{escaped}"')
            else:
                lines.append(f"{key}: {value}")
    lines.append("---")
    return lines


def emit_scales(scales: List[Dict[str, str]]) -> List[str]:
    by_scale: Dict[str, List[Dict[str, str]]] = {}
    for row in scales:
        code = row.get("scale_code", "")
        if code:
            by_scale.setdefault(code, []).append(row)
    lines: List[str] = []
    for scale_code in sorted(by_scale.keys()):
        rows = sorted(by_scale[scale_code], key=sort_key)
        scale_type = rows[0].get("type", "single") or "single"
        lines.extend(["", f"## Escala: {scale_code}", f"type: {scale_type}"])
        for row in rows:
            option_code = row.get("option_code", "")
            option_text = row.get("option_text", "")
            if option_code and option_text:
                lines.append(f"- {option_code} | {option_text}")
    return lines


def group_subquestions(subquestions: List[Dict[str, str]]) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = {}
    for row in subquestions:
        qcode = row.get("question_code", "")
        if qcode:
            grouped.setdefault(qcode, []).append(row)
    for qcode in grouped:
        grouped[qcode].sort(key=sort_key)
    return grouped


def group_options(options: List[Dict[str, str]]) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = {}
    for row in options:
        qcode = row.get("question_code", "")
        if qcode:
            grouped.setdefault(qcode, []).append(row)
    for qcode in grouped:
        grouped[qcode].sort(key=sort_key)
    return grouped


def subgroup_labels(subgroups: List[Dict[str, str]]) -> Dict[tuple, str]:
    labels: Dict[tuple, str] = {}
    for row in subgroups:
        group_code = row.get("group_code", "")
        subgroup_code = row.get("subgroup_code", "")
        title = row.get("subgroup_title", "") or row.get("title", "")
        if group_code and subgroup_code:
            labels[(group_code, subgroup_code)] = " ".join(part for part in [subgroup_code, title] if part).strip()
    return labels


def emit_kv(lines: List[str], key: str, value: str) -> None:
    value = clean(value)
    if value:
        lines.append(f"{key}: {value}")


def emit_question_options(lines: List[str], question: Dict[str, str]) -> None:
    for key in QUESTION_OPTION_KEYS:
        value = question.get(key, "")
        if value:
            emit_kv(lines, key, value)


def emit_question_body(lines: List[str], text: str) -> None:
    text = clean(text)
    if text:
        lines.append("")
        lines.extend(text.splitlines())


def emit_inline_options(lines: List[str], options: List[Dict[str, str]]) -> None:
    if not options:
        return
    lines.extend(["", "options:"])
    for row in options:
        code = row.get("option_code", "")
        text = row.get("option_text", "")
        if code and text:
            lines.append(f"- {code} | {text}")


def emit_subquestions(lines: List[str], rows: List[Dict[str, str]], section_name: str = "subquestions") -> None:
    filtered = [r for r in rows if r.get("code") and r.get("text")]
    if not filtered:
        return
    lines.extend(["", f"{section_name}:"])
    for row in filtered:
        code = row.get("code", "")
        text = row.get("text", "")
        visible_if = row.get("visible_if", "")
        attrs = []
        if visible_if:
            attrs.append(f"visible_if={visible_if}")
        suffix = f" {{{'; '.join(attrs)}}}" if attrs else ""
        lines.append(f"- {code} | {text}{suffix}")


def emit_subquestion_evidence_questions(lines: List[str], parent_question: Dict[str, str], rows: List[Dict[str, str]]) -> None:
    parent_code = parent_question.get("code", "")
    parent_type = parent_question.get("type", "")
    if not parent_code:
        return
    for row in rows:
        evidence_type = row.get("evidence", "")
        sub_code = row.get("code", "")
        if not evidence_type or not sub_code:
            continue
        if parent_type == "adoption":
            source_code = f"{parent_code}ext"
        else:
            source_code = parent_code
        evidence_code = f"{source_code}{sub_code}_evi"
        lines.extend(["", f"### {evidence_code} [{evidence_type}]"])
        mandatory = row.get("evidence_mandatory", "") or "false"
        lines.append(f"mandatory: {mandatory}")
        lines.append(f"visible_if: {source_code}.{sub_code} == Y")
        if row.get("evidence_allowed_filetypes", ""):
            lines.append(f"allowed_filetypes: {row['evidence_allowed_filetypes']}")
        if row.get("evidence_min_files", ""):
            lines.append(f"min_files: {row['evidence_min_files']}")
        if row.get("evidence_max_files", ""):
            lines.append(f"max_files: {row['evidence_max_files']}")
        text = row.get("evidence_text", "") or f"Anexe evidência documental referente ao item {sub_code}."
        lines.extend(["", text])


def emit_group(lines: List[str], group: Dict[str, str]) -> None:
    code = group.get("group_code", "") or group.get("code", "")
    name = group.get("group_name", "") or group.get("name", "")
    if not code:
        return
    lines.extend(["", f"## Grupo: {code} | {name}"])
    desc = group.get("description", "")
    if desc:
        lines.append(f"> {desc}")


def emit_questions_for_group(
    lines: List[str],
    questions: Iterable[Dict[str, str]],
    subquestions_by_q: Dict[str, List[Dict[str, str]]],
    options_by_q: Dict[str, List[Dict[str, str]]],
    subgroup_by_key: Dict[tuple, str],
) -> None:
    emitted_subgroups = set()
    for q in sorted(questions, key=sort_key):
        code = q.get("code", "")
        qtype = q.get("type", "")
        if not code or not qtype:
            continue
        lines.extend(["", f"### {code} [{qtype}]"])
        emit_question_options(lines, q)
        group_code = q.get("group_code", "")
        subgroup_code = q.get("subgroup_code", "")
        if subgroup_code and subgroup_code not in emitted_subgroups:
            subgroup_key = (group_code, subgroup_code)
            if subgroup_key not in subgroup_by_key:
                raise ValueError(f"Questão {code} referencia subgrupo inexistente: {subgroup_code}")
            emit_kv(lines, "subgroup", subgroup_by_key[subgroup_key])
            emitted_subgroups.add(subgroup_code)
        emit_question_body(lines, q.get("text", ""))
        q_subquestions = subquestions_by_q.get(code, [])
        if qtype == "adoption":
            emit_subquestions(lines, q_subquestions, section_name="detail_options")
        else:
            emit_inline_options(lines, options_by_q.get(code, []))
            emit_subquestions(lines, q_subquestions, section_name="subquestions")
        emit_subquestion_evidence_questions(lines, q, q_subquestions)


def convert_xlsx_to_md(input_path: Path, org: Optional[str] = None) -> str:
    wb = load_workbook(input_path, data_only=False)
    meta = meta_for_org(survey_meta(wb), org)
    groups = sorted(sheet_rows(wb, "groups"), key=sort_key)
    scales = sheet_rows(wb, "scales")
    questions = sheet_rows(wb, "questions")
    subquestions = sheet_rows(wb, "subquestions")
    options = sheet_rows(wb, "options")
    subgroups = sheet_rows(wb, "subgroups")
    original_question_codes = {row.get("code", "") for row in questions if row.get("code", "")}
    questions, subquestions, options = filter_rows_for_org(questions, subquestions, options, org)
    if org:
        validate_visible_if_references(questions, org, original_question_codes)

    subquestions_by_q = group_subquestions(subquestions)
    options_by_q = group_options(options)
    subgroup_by_key = subgroup_labels(subgroups)
    questions_by_group: Dict[str, List[Dict[str, str]]] = {}
    for q in questions:
        group_code = q.get("group_code", "")
        questions_by_group.setdefault(group_code, []).append(q)

    title = meta.get("title", "Questionário")
    lines = emit_frontmatter(meta)
    lines.extend(["", f"# {title}"])
    lines.extend(emit_scales(scales))

    declared_groups = {g.get("group_code", "") or g.get("code", "") for g in groups}
    for group in groups:
        gcode = group.get("group_code", "") or group.get("code", "")
        if org and not questions_by_group.get(gcode, []):
            continue
        emit_group(lines, group)
        emit_questions_for_group(lines, questions_by_group.get(gcode, []), subquestions_by_q, options_by_q, subgroup_by_key)

    for group_code in sorted(k for k in questions_by_group.keys() if k not in declared_groups):
        lines.extend(["", f"## Grupo: {group_code} | {group_code}"])
        emit_questions_for_group(lines, questions_by_group[group_code], subquestions_by_q, options_by_q, subgroup_by_key)

    return "\n".join(lines).rstrip() + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description="Converte SurveyXLSX para SurveyMD")
    parser.add_argument("input", type=Path, help="Arquivo .xlsx de entrada")
    parser.add_argument("output", type=Path, help="Arquivo .md de saída")
    parser.add_argument("--org", default=None, help="Gera apenas a variante da organizacao informada")
    args = parser.parse_args()
    md = convert_xlsx_to_md(args.input, org=args.org)
    args.output.write_text(md, encoding="utf-8")
    print(f"Arquivo Markdown gerado: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
